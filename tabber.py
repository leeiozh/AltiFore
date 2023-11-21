import openpyxl  # conda install -c conda-forge openpyxl
from openpyxl.styles import PatternFill
import pandas as pd  # conda install -c conda-forge pandas
import datetime as dt
from calculator import calc_sunset_sunrise
from const import SAT_NAMES, color_rgd


def convert_list(name, sheet, res_list):
    st = 0
    while len(res_list[st]) == 0:
        st += 1
    curr_date = res_list[st][0]['date']

    sr, ss = res_list_to_sr_ss(res_list[st])
    sheet.loc[sheet['date'] == curr_date, "sunrise"] = sr
    sheet.loc[sheet['date'] == curr_date, "sunset"] = ss

    curr_flight = 0
    for i in range(st, len(res_list)):
        if len(res_list[i]) != 0:
            if curr_date != res_list[i][0]['date']:
                curr_date = res_list[i][0]['date']
                curr_flight = 0

                sr, ss = res_list_to_sr_ss(res_list[i])
                sheet.loc[sheet['date'] == curr_date, "sunrise"] = sr
                sheet.loc[sheet['date'] == curr_date, "sunset"] = ss

            for j in range(len(res_list[i])):
                curr_flight += 1
                sheet.loc[sheet['date'] == curr_date, f"flight {curr_flight}"] = res_list[i][j]['time']

    sheet.dropna(axis=1, how='all', inplace=True)
    sheet = pd.concat([sheet, pd.DataFrame({'name': [sat_name for sat_name in SAT_NAMES]})], axis=1)

    sheet.to_excel(name, index=False, engine='openpyxl')
    wb = openpyxl.load_workbook(name)
    ws = wb['Sheet1']
    alphabet = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T']

    for s in range(len(SAT_NAMES)):
        ws[str(alphabet[sheet.shape[1] - 1] + str(s + 2))].fill = PatternFill(patternType='solid', fgColor=color_rgd[s])

    letter = 2
    # filling a color
    try:
        row = (sheet.index[sheet['date'] == curr_date][0]) + 2
        for i in range(len(res_list)):
            if len(res_list[i]) != 0:
                if curr_date != res_list[i][0]['date']:
                    curr_date = res_list[i][0]['date']
                    letter = 2
                    row = (sheet.index[sheet['date'] == curr_date][0]) + 2

                for res_it in res_list[i]:
                    letter += 1
                    ws[str(alphabet[letter] + str(row))].fill = PatternFill(patternType='solid',
                                                                            fgColor=res_it['color'])
    except:
        pass

    wb.save(name)


def prepare_sheet(start, nums):
    res = pd.DataFrame(
        columns=['date', 'sunrise', 'sunset', 'flight 1', 'flight 2', 'flight 3', 'flight 4', 'flight 5', 'flight 6',
                 'flight 7', 'flight 8', 'flight 9', 'flight 10', 'flight 11', 'flight 12',
                 'flight 13', 'flight 14', 'flight 15', 'flight 16'])
    res['date'] = [(start + dt.timedelta(days=i)).strftime("%d/%m/%Y") for i in range(nums)]
    return res


def res_list_to_sr_ss(inp):
    num_day = dt.datetime.strptime(inp[0]['date'], "%d/%m/%Y").timetuple()[7]
    # sunrise for first flight
    sr, _ = calc_sunset_sunrise(inp[0]['lat'], inp[0]['lon'], num_day)
    sr, _ = sr_ss(sr, _)
    # sunset for last flight
    _, ss = calc_sunset_sunrise(inp[-1]['lat'], inp[-1]['lon'], num_day)
    _, ss = sr_ss(_, ss)
    return sr, ss


def sr_ss(sr, ss):
    if sr == -1:
        if ss == 1:
            return "PD", "PD"
        else:
            return "PN", "PN"
    else:
        sr_h = int(sr)
        sr_m = int(60 * (sr - sr_h))
        if sr_h < 10:
            sr_h = "0" + str(sr_h)
        else:
            sr_h = str(sr_h)
        if sr_m < 10:
            sr_m = "0" + str(sr_m)
        else:
            sr_m = str(sr_m)
        ss_h = int(ss)
        ss_m = int(60 * (ss - ss_h))
        if ss_h < 10:
            ss_h = "0" + str(ss_h)
        else:
            ss_h = str(ss_h)
        if ss_m < 10:
            ss_m = "0" + str(ss_m)
        else:
            ss_m = str(ss_m)
    return sr_h + ":" + sr_m, ss_h + ":" + ss_m
